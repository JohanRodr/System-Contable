import os
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
import calendar

def export_to_excel(template_path_compras, template_path_ventas, usuario=None, year=None, month=None, tipo_libro=None):
    try:
        # Verificar si los archivos de plantilla existen según el tipo de libro
        if tipo_libro in ["LIBRO DE COMPRAS", "AMBOS"]:
            print(f"Verificando la existencia del archivo de plantilla de compras: {template_path_compras}")
            if not os.path.exists(template_path_compras):
                print(f"El archivo de plantilla de compras no existe: {template_path_compras}")
                return False
        if tipo_libro in ["LIBRO DE VENTAS", "AMBOS"]:
            print(f"Verificando la existencia del archivo de plantilla de ventas: {template_path_ventas}")
            if not os.path.exists(template_path_ventas):
                print(f"El archivo de plantilla de ventas no existe: {template_path_ventas}")
                return False

        # Cargar datos de los contribuyentes para obtener el RIF del usuario
        with open(os.path.join(os.path.dirname(__file__), "contribuyentes.json"), "r", encoding="utf-8") as file:
            contribuyentes = json.load(file)
            usuario_data = next((user for user in contribuyentes if user.get("Nombre:").upper() == usuario.upper() and user.get("Tipo de Empresa:") == "Empresa"), None)
            if not usuario_data:
                print(f"Usuario {usuario} no encontrado en contribuyentes.json")
                return False
            rif_usuario = usuario_data.get("Número R.I.F.:", "")

        # Cargar datos de las facturas
        with open(os.path.join(os.path.dirname(__file__), "datos_guardados.json"), "r", encoding="utf-8") as file:
            facturas = json.load(file)

        # Filtrar las facturas por usuario si se proporciona
        if usuario:
            facturas = [factura for factura in facturas if factura.get("Empresa").upper() == usuario.upper()]

        # Filtrar las facturas por año y mes si se proporcionan
        if year and month:
            facturas = [factura for factura in facturas if datetime.strptime(factura['Fecha'], '%Y-%m-%d').year == year and datetime.strptime(factura['Fecha'], '%Y-%m-%d').month == month]

        # Ordenar las facturas por fecha
        facturas.sort(key=lambda x: datetime.strptime(x['Fecha'], '%Y-%m-%d'))

        # Filtrar facturas por tipo de libro
        facturas_compras = [factura for factura in facturas if factura.get("Tipo de Transacción") == "Compras"]
        facturas_ventas = [factura for factura in facturas if factura.get("Tipo de Transacción") == "Ventas"]

        # Crear la estructura de carpetas relativa a la ubicación del script
        script_dir = os.path.dirname(__file__)
        base_path = os.path.join(script_dir, "INFORMES CONTADORES", "EMPRESAS", usuario.upper(), "LIBROS DE COMPRAS Y VENTAS", str(year), f"{month:02d}")
        os.makedirs(base_path, exist_ok=True)

        if tipo_libro in ["LIBRO DE COMPRAS", "AMBOS"]:
            if not export_compras(template_path_compras, usuario, rif_usuario, year, month, facturas_compras, base_path):
                return False

        if tipo_libro in ["LIBRO DE VENTAS", "AMBOS"]:
            if not export_ventas(template_path_ventas, usuario, rif_usuario, year, month, facturas_ventas, base_path):
                return False

        return True

    except Exception as e:
        print(f"Error al exportar datos: {str(e)}")
        return False

def export_compras(template_path, usuario, rif_usuario, year, month, facturas, base_path):
    try:
        if not facturas:
            print("No hay facturas de compras para exportar.")
            return False

        # Cargar el archivo Excel existente
        wb = load_workbook(template_path)
        ws = wb.active

        # Modificar la casilla D-1 con el nombre del usuario en mayúsculas
        ws['D1'] = usuario.upper()

        # Modificar la casilla D-2 con el RIF del usuario
        ws['D2'] = rif_usuario

        # Modificar la casilla D-4 con el rango de fechas correspondiente
        first_day = f"01/{month:02d}/{year}"
        last_day = f"{calendar.monthrange(year, month)[1]}/{month:02d}/{year}"
        ws['D4'] = f"{first_day} al {last_day}"

        # Copiar la fila 10 para mantener el formato
        row_10 = [cell.value for cell in ws[10]]
        row_10_styles = [cell._style for cell in ws[10]]

        # Preparar los datos para el Excel
        for i, factura in enumerate(facturas, start=10):  # Comenzar desde la fila 10
            row_index = i  # Ajustar el índice de fila para la nueva fila

            # Insertar una nueva fila para cada factura
            ws.insert_rows(row_index)

            # Pegar el formato copiado
            for col_index, (value, style) in enumerate(zip(row_10, row_10_styles), start=1):
                cell = ws.cell(row=row_index, column=col_index)
                cell.value = value
                cell._style = style

            # Insertar los datos de la factura
            ws[f'A{row_index}'] = row_index - 9  # Columna A, comenzando desde 1
            ws[f'B{row_index}'] = datetime.strptime(factura['Fecha'], '%Y-%m-%d').strftime('%d/%m/%Y')
            ws[f'C{row_index}'] = factura['RIF']
            ws[f'D{row_index}'] = factura['Nombre del Cliente'].upper()
            ws[f'L{row_index}'] = factura['Número de Documento']
            ws[f'M{row_index}'] = factura['Número de Control']
            ws[f'AB{row_index}'] = float(factura['Base imponible 16%']) if factura['Base imponible 16%'] else 0.0
            ws[f'AC{row_index}'] = 16  # Colocar el valor 16 en la columna AC de cada fila
            ws[f'AD{row_index}'] = f"=AB{row_index}*16%"  # Colocar la fórmula en la columna AD de la fila actual
            ws[f'W{row_index}'] = f"=AB{row_index}+AD{row_index}"  # Colocar la fórmula en la columna W de la fila actual

        # Calcular la fila de la última factura
        last_row = len(facturas) + 9

        # Eliminar la fila adicional si existe
        if ws.cell(row=last_row + 1, column=1).value is None:
            ws.delete_rows(last_row + 1)

        # Colocar el valor 16 en la columna AC de la última fila
        ws[f'AC{last_row}'] = 16

        # Actualizar las fórmulas de suma en las columnas de R a AF, excepto AC
        for col in range(18, 33):  # Columnas R (18) a AF (33)
            if col == 29:  # Columna AC (29)
                continue
            col_letter = get_column_letter(col)
            sum_formula = f"=SUM({col_letter}10:{col_letter}{last_row})"
            ws.cell(row=last_row + 1, column=col, value=sum_formula)

        # Calcular las posiciones dinámicas para las fórmulas específicas
        base_row = 20  # La fila base para las fórmulas
        h21_row = base_row + len(facturas) - 1
        i21_row = h21_row
        h24_row = h21_row + 3
        i28_row = h21_row + 7
        i29_row = h21_row + 8
        i31_row = h21_row + 10

        # Colocar las fórmulas específicas en las celdas indicadas
        ws[f'H{h21_row}'] = f"=+AB{last_row + 1}"
        ws[f'I{i21_row}'] = f"=+AD{last_row + 1}"
        ws[f'I{i28_row}'] = f"=+H{h21_row}+H{h24_row}+I{i21_row}"
        ws[f'I{i29_row}'] = f"=+I{i21_row}"
        ws[f'I{i31_row}'] = f"=+I{i21_row}"

        # Aplicar formato de número con separador de miles y decimales y establecer el tamaño de la fuente
        number_format = '#,##0.00'
        font_size_8 = Font(size=8)
        for row in range(10, last_row + 1):
            for col in range(1, 33):  # Columnas A (1) a AF (33)
                cell = ws.cell(row=row, column=col)
                if col != 4:  # Excluir la columna D (4) que contiene texto
                    cell.number_format = number_format
                cell.font = font_size_8

        # Asegurar que la columna A tenga el formato general
        for row in range(10, last_row + 1):
            ws[f'A{row}'].number_format = 'General'

        # Aplicar bordes superiores a las celdas desde la fila 11, columna R hasta la columna AF
        thin_border = Border(top=Side(style='thin'))
        for col in range(18, 33):  # Columnas R (18) a AF (33)
            cell = ws.cell(row=last_row + 1, column=col)
            cell.border = thin_border

        # Guardar el archivo Excel con los datos agregados
        month_name = datetime(year, month, 1).strftime('%B').upper()
        excel_path = os.path.join(base_path, f'LIBRO DE COMPRAS {month_name} {year} - {usuario.upper()}.xlsx')
        wb.save(excel_path)

        print(f"Datos exportados exitosamente a {excel_path}")
        return True

    except Exception as e:
        print(f"Error al exportar datos de compras: {str(e)}")
        return False

def export_ventas(template_path, usuario, rif_usuario, year, month, facturas, base_path):
    try:
        if not facturas:
            print("No hay facturas de ventas para exportar.")
            return False

        # Cargar el archivo Excel existente
        wb = load_workbook(template_path)
        ws = wb.active

        # Modificar la casilla D-1 con el nombre del usuario en mayúsculas
        ws['D1'] = usuario.upper()

        # Modificar la casilla D-2 con el RIF del usuario
        ws['D2'] = rif_usuario

        # Modificar la casilla D-4 con el rango de fechas correspondiente
        first_day = f"01/{month:02d}/{year}"
        last_day = f"{calendar.monthrange(year, month)[1]}/{month:02d}/{year}"
        ws['D4'] = f"{first_day} al {last_day}"

        # Preparar los datos para el Excel
        for i, factura in enumerate(facturas, start=10):  # Comenzar desde la fila 10
            row_index = i  # Ajustar el índice de fila para la nueva fila

            # Insertar una nueva fila para cada factura
            ws.insert_rows(row_index)

            # Insertar los datos de la factura
            ws[f'A{row_index}'] = row_index - 9  # Columna A, comenzando desde 1
            ws[f'B{row_index}'] = datetime.strptime(factura['Fecha'], '%Y-%m-%d').strftime('%d/%m/%Y')
            ws[f'C{row_index}'] = factura['RIF']
            ws[f'D{row_index}'] = factura['Nombre del Cliente'].upper()
            ws[f'F{row_index}'] = factura['Número de Documento']
            ws[f'H{row_index}'] = factura['Número de Control']
            ws[f'K{row_index}'] = "01 Registro"  # Colocar "01 Registro" en la columna K
            ws[f'R{row_index}'] = float(factura['Base imponible 16%']) if factura['Base imponible 16%'] else 0.0
            ws[f'S{row_index}'] = 16  # Colocar el valor 16 en la columna S de cada fila
            ws[f'T{row_index}'] = f"=R{row_index}*16%"  # Colocar la fórmula en la columna T de la fila actual

            # Colocar 0 en las columnas P-Q, U-W
            for col in ['P', 'Q', 'U', 'V', 'W']:
                ws[f'{col}{row_index}'] = 0

            # Colocar 16 en la columna X
            ws[f'X{row_index}'] = 16

            # Colocar 0 en las columnas Y-AA
            for col in ['Y', 'Z', 'AA']:
                ws[f'{col}{row_index}'] = 0

        # Calcular la fila de la última factura
        last_row = len(facturas) + 9

        # Eliminar la fila adicional si existe
        if ws.cell(row=last_row + 1, column=1).value is None:
            ws.delete_rows(last_row + 1)

        # Colocar el valor 16 en la columna S de la última fila
        ws[f'S{last_row}'] = 16

        # Actualizar las fórmulas de suma en las columnas de O a R, T a W, y Y a AA
        for col in list(range(15, 19)) + list(range(20, 24)) + list(range(25, 28)):  # Columnas O (15) a R (18), T (20) a W (23), y Y (25) a AA (27)
            col_letter = get_column_letter(col)
            sum_formula = f"=SUM({col_letter}10:{col_letter}{last_row})"
            ws.cell(row=last_row + 1, column=col, value=sum_formula)

        # Colocar las fórmulas en la columna O
        for i in range(10, last_row + 1):
            ws[f'O{i}'] = f"=+R{i}+T{i}"

        # Colocar la fórmula de suma en la última fila de la columna O
        ws[f'O{last_row + 1}'] = f"=SUM(O10:O{last_row})"

        # Colocar las fórmulas en las columnas E y F según el número de facturas, sumando una fila más
        if len(facturas) == 1:
            ws[f'E18'] = f"=+W11+R11"
            ws[f'F18'] = f"=+Y11+T11"
        elif len(facturas) == 2:
            ws[f'E19'] = f"=+W12+R12"
            ws[f'F19'] = f"=+Y12+T12"
        elif len(facturas) >= 3:
            ws[f'E20'] = f"=+W13+R13"
            ws[f'F20'] = f"=+Y13+T13"

        # Aplicar formato de número con separador de miles y decimales y establecer el tamaño de la fuente
        number_format = '#,##0.00'
        font_size_8 = Font(size=8)
        alignment_center = Alignment(horizontal='center', vertical='center')
        for row in range(10, last_row + 1):
            for col in range(1, 28):  # Columnas A (1) a AA (27)
                cell = ws.cell(row=row, column=col)
                if col != 4:  # Excluir la columna D (4) que contiene texto
                    cell.number_format = number_format
                cell.font = font_size_8
                if col in [6, 8]:  # Centrar solo las columnas F (6) y H (8)
                    cell.alignment = alignment_center

        # Asegurar que la columna A tenga el formato general
        for row in range(10, last_row + 1):
            ws[f'A{row}'].number_format = 'General'

        # Aplicar bordes superiores a las celdas desde la fila 11, columna R hasta la columna AF
        thin_border = Border(top=Side(style='thin'))
        for col in range(18, 33):  # Columnas R (18) a AF (33)
            cell = ws.cell(row=last_row + 1, column=col)
            cell.border = thin_border

        # Guardar el archivo Excel con los datos agregados
        month_name = datetime(year, month, 1).strftime('%B').upper()
        excel_path = os.path.join(base_path, f'LIBRO DE VENTAS {month_name} {year} - {usuario.upper()}.xlsx')
        wb.save(excel_path)

        print(f"Datos exportados exitosamente a {excel_path}")
        return True

    except Exception as e:
        print(f"Error al exportar datos de ventas: {str(e)}")
        return False

# Ejemplo de llamada a la función desde otro lugar en el código
def some_other_function():
    usuario = "Nombre de la Empresa"
    year = 2025
    month = 1
    tipo_libro = "AMBOS"
    template_path_compras = os.path.join(os.path.dirname(__file__), "Template/LIBRO DE COMPRAS.xlsx")
    template_path_ventas = os.path.join(os.path.dirname(__file__), "Template/LIBRO DE VENTAS.xlsx")
    export_to_excel(template_path_compras, template_path_ventas, usuario=usuario, year=year, month=month, tipo_libro=tipo_libro)

if __name__ == "__main__":
    some_other_function()
